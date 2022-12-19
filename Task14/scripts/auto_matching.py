import openpyxl
import csv
import os
import string
import re
from openpyxl import load_workbook
from py_crunchbase import PyCrunchbase, Collections
import urllib3
import requests
import simplejson
import socket
from time import sleep

# # PC
# # INC
# # LLC
# # USA
# # corp
# # ltd
# # pllc
# # corporation
# # &amp
# # corp
# # assn
# # co
# del_list = [" pc",
#             "pc.",
#             " inc",
#             "inc.",
#             " llc",
#             "llc.",
#             "l.l.c.",
#             " usa",
#             "usa.",
#             " corp",
#             "corp.",
#             " ltd",
#             "ltd.",
#             " pllc",
#             "pllc.",
#             "corporation",
#             "&amp",
#             " assn",
#             "assn.",
#             "co.",
#             ",",
#             "."]

# def f_process(original_name):
#     res = original_name.lower()
#     for del_word in del_list:
#         res = res.replace(del_word, "")

#     res = re.sub(r'\(.*\)', "", res)  
#     return res

MY_KEY = "ad5244ad6326a6f13b8ccce0c7dd6e41"
res_path = "./gen_4/SBIR_remains.csv"
xlsx_path = "./original1/"
csv_path_pre = "./gen_4/SBIR_"
LIMIT = 5
SLEEP_TIME = 15
MOD_NUM = 100

remains = []


head = ["origin_name", "domain"]
for i in range(0, LIMIT):
    head.append("permalink_" + str(i))
    head.append("uuid_" + str(i))
remains.append(head)




def xlsx2csv(xlsx_path, csv_path_pre):
    wb = load_workbook(filename = xlsx_path)
    ws = wb.active
    csv_data = []
    
    csv_data.append(head)

    csv_cnt = 0
    row_cnt = 1
    for val in ws.iter_rows(values_only = True):
        cur_name = val[1]
        cur_domain = val[12]
        if cur_name == "Company Name":
            continue
        
        #processed_name = f_process(cur_name)
        cur_row = [cur_name, cur_domain]
        if cur_name == "":
            continue
        try:
            API_KEY = MY_KEY
            pycb = PyCrunchbase(API_KEY)
            api = pycb.autocomplete_api()
            for entity in api.autocomplete(cur_name).limit(LIMIT).select_collections(Collections.Organizations.companies).execute():
                cur_row.append(entity.permalink)
                cur_row.append(entity.uuid)
        except (simplejson.errors.JSONDecodeError,\
                 requests.exceptions.JSONDecodeError,\
                 requests.exceptions.ConnectionError,\
                 urllib3.exceptions.ProtocolError,\
                 urllib3.exceptions.MaxRetryError,\
                 urllib3.exceptions.NewConnectionError,\
                 socket.gaierror):
            sleep(SLEEP_TIME)
            print(cur_row)
            csv_data.append(cur_row)
            continue
        
        print(len(csv_data))
        print(cur_row)
        csv_data.append(cur_row)
        row_cnt += 1
        if row_cnt % MOD_NUM == 0:
            cur_csv_name = csv_path_pre + str(csv_cnt) + ".csv"

            with open(cur_csv_name, 'w') as csv_file:
                writer = csv.writer(csv_file, delimiter = ',')
                for line in csv_data:
                    writer.writerow(line)

            csv_cnt += 1
            print(cur_csv_name + " success!")
            csv_data = []
            csv_data.append(head)
            row_cnt += 1
    
    # remains
    for idx in range(1, len(csv_data)):
        remains.append(csv_data[idx]) 


def mod1000():
    all_xlsx_path = xlsx_path
    all_csv_path_pre = csv_path_pre

    for root, dirs, files in os.walk(all_xlsx_path):
        for file_name in files:
            file_name_list = file_name.split('.')
            cur_xlsx_path = all_xlsx_path + file_name
            cur_csv_path_pre = all_csv_path_pre + file_name_list[0] + "_"
            xlsx2csv(cur_xlsx_path, cur_csv_path_pre)

def main():
    mod1000()
    remain_csv_path_pre = res_path
    with open(remain_csv_path_pre, 'w') as csv_file:
        writer = csv.writer(csv_file, delimiter = ',')
        for line in remains:
            writer.writerow(line)
    print(remain_csv_path_pre + " success!")

main()
        

